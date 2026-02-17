#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CIS Benchmark Kural Bağımlılıkları Excel Oluşturucu
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# Workbook oluştur
wb = Workbook()

# Stiller
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_fill_green = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
header_fill_orange = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
header_fill_purple = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
wrap_alignment = Alignment(wrap_text=True, vertical='top')

def apply_header_style(ws, row, fill=header_fill):
    for cell in ws[row]:
        cell.font = header_font
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def apply_data_style(ws, start_row, end_row):
    for row in ws.iter_rows(min_row=start_row, max_row=end_row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = wrap_alignment

def auto_column_width(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

# ============================================
# SHEET 1: Ana Bağımlılık Tablosu
# ============================================
ws1 = wb.active
ws1.title = "Ana Bagimlilik Tablosu"

headers1 = ["Kural ID", "Kural Adi", "Bagimlilik Tipi", "Bagli Oldugu Kural(lar)", "Kosul", "Aciklama"]
ws1.append(headers1)
apply_header_style(ws1, 1)

data1 = [
    # FIREWALL - UFW
    ["", "FIREWALL - UFW ZİNCİRİ", "", "", "", ""],
    ["4.2.1", "ufw is installed", "Seçim Bağımlılığı", "4.1.1", "firewall_choice = ufw", "UFW seçildiyse uygulanır"],
    ["4.2.2", "iptables-persistent not installed", "Sıralı", "4.2.1", "ufw kurulu", "UFW ile çakışan paket kaldırılmalı"],
    ["4.2.3", "ufw service is enabled", "Sıralı", "4.2.1", "ufw kurulu", "Önce kurulum, sonra servis"],
    ["4.2.4", "ufw loopback traffic configured", "Sıralı", "4.2.3", "ufw servisi aktif", "Servis aktif olmalı ki kural eklensin"],
    ["4.2.5", "ufw outbound connections", "Sıralı", "4.2.3", "ufw servisi aktif", "Servis aktif olmalı"],
    ["4.2.6", "ufw firewall rules exist", "Sıralı", "4.2.3", "ufw servisi aktif", "Servis aktif olmalı"],
    ["4.2.7", "ufw default deny policy", "Sıralı", "4.2.4, 4.2.5, 4.2.6", "Diğer kurallar önce", "Default deny en son, yoksa kilitlenme"],

    # FIREWALL - NFTABLES
    ["", "FIREWALL - NFTABLES ZİNCİRİ", "", "", "", ""],
    ["4.3.2", "ufw disabled with nftables", "Seçim Bağımlılığı", "4.1.1", "firewall_choice = nftables", "nftables seçildiyse UFW kapatılmalı"],
    ["4.3.7", "nftables connections configured", "Sıralı", "4.3.2", "nftables aktif", "Önce servis, sonra kurallar"],
    ["4.3.8", "nftables default deny", "Sıralı", "4.3.7", "Bağlantı kuralları önce", "Default deny en son uygulanmalı"],
    ["4.3.9", "nftables service enabled", "Sıralı", "4.3.8", "Kurallar hazır", "Kurallar hazırlandıktan sonra servis"],

    # FIREWALL - IPTABLES
    ["", "FIREWALL - IPTABLES ZİNCİRİ", "", "", "", ""],
    ["4.4.1.x", "iptables rules", "Seçim Bağımlılığı", "4.1.1", "firewall_choice = iptables", "Legacy sistemler için"],

    # KARŞILIKLI DIŞLAMA
    ["", "KARŞILIKLI DIŞLAMA (MUTUAL EXCLUSION)", "", "", "", ""],
    ["4.2.x", "UFW kuralları", "Karşılıklı Dışlama", "-", "4.3.x ve 4.4.x ile birlikte UYGULANMAZ", "Sadece biri seçilmeli"],
    ["4.3.x", "nftables kuralları", "Karşılıklı Dışlama", "-", "4.2.x ve 4.4.x ile birlikte UYGULANMAZ", "Sadece biri seçilmeli"],
    ["4.4.x", "iptables kuralları", "Karşılıklı Dışlama", "-", "4.2.x ve 4.3.x ile birlikte UYGULANMAZ", "Sadece biri seçilmeli"],

    # GDM ZİNCİRİ
    ["", "GDM ZİNCİRİ", "", "", "", ""],
    ["1.7.1", "GDM is removed", "Koşullu", "-", "profile = Level 2 Server", "Sadece sunucu ortamında"],
    ["1.7.2", "GDM login banner", "Koşullu", "1.7.1 (tersi)", "gdm3 kurulu", "GDM kuruluysa uygulanır"],
    ["1.7.3", "GDM disable-user-list", "Koşullu", "1.7.1 (tersi)", "gdm3 kurulu", "GDM kuruluysa uygulanır"],
    ["1.7.4", "GDM screen locks idle", "Koşullu", "1.7.1 (tersi)", "gdm3 + gnome kurulu", "GNOME masaüstü gerekli"],
    ["1.7.5", "GDM screen locks override", "Koşullu", "1.7.4", "gdm3 kurulu", "1.7.4 ile birlikte"],
    ["1.7.6", "GDM automatic mounting disabled", "Koşullu", "1.7.1 (tersi)", "gdm3 kurulu", "GDM kuruluysa uygulanır"],
    ["1.7.7", "GDM autorun-never enabled", "Koşullu", "1.7.1 (tersi)", "gdm3 kurulu", "GDM kuruluysa uygulanır"],
    ["1.7.8", "GDM autorun-never override", "Koşullu", "1.7.7", "gdm3 kurulu", "1.7.7 ile birlikte"],
    ["1.7.9", "GDM disable-print-setup", "Koşullu", "1.7.1 (tersi)", "gdm3 kurulu", "GDM kuruluysa uygulanır"],
    ["1.7.10", "GDM XDMCP disabled", "Koşullu", "1.7.1 (tersi)", "gdm3 kurulu", "GDM kuruluysa uygulanır"],

    # PAM ZİNCİRİ
    ["", "PAM ZİNCİRİ", "", "", "", ""],
    ["5.3.1.1", "PAM latest version", "Ön Koşul", "-", "-", "Tüm PAM kurallarının temeli"],
    ["5.3.2.1", "pam_unix enabled", "Sıralı", "5.3.1.1", "PAM ≥1.5.3-5", "PAM güncel olmalı"],
    ["5.3.2.2", "pam_faillock enabled", "Sıralı", "5.3.1.1", "PAM ≥1.5.3-5", "PAM güncel olmalı"],
    ["5.3.2.3", "pam_pwquality enabled", "Sıralı", "5.3.1.1", "PAM ≥1.5.3-5", "PAM güncel olmalı"],
    ["5.3.2.4", "pam_pwhistory enabled", "Sıralı", "5.3.1.1", "PAM ≥1.5.3-5", "PAM güncel olmalı"],
    ["5.3.3.1.1", "faillock attempts", "Sıralı", "5.3.2.2", "pam_faillock etkin", "Modül etkin olmalı"],
    ["5.3.3.1.2", "faillock unlock time", "Sıralı", "5.3.2.2", "pam_faillock etkin", "Modül etkin olmalı"],
    ["5.3.3.1.3", "faillock root lockout", "Sıralı", "5.3.2.2", "pam_faillock etkin", "Modül etkin olmalı"],
    ["5.3.3.2.x", "pwquality settings", "Sıralı", "5.3.2.3", "pam_pwquality etkin", "Modül etkin olmalı"],
    ["5.3.3.3.x", "pwhistory settings", "Sıralı", "5.3.2.4", "pam_pwhistory etkin", "Modül etkin olmalı"],
    ["5.3.3.4.x", "pam_unix settings", "Sıralı", "5.3.2.1", "pam_unix etkin", "Modül etkin olmalı"],

    # SSH ZİNCİRİ
    ["", "SSH ZİNCİRİ", "", "", "", ""],
    ["5.1.1", "sshd_config permissions", "Paket Bağımlılığı", "-", "openssh-server kurulu", "SSH kurulu olmalı"],
    ["5.1.2", "SSH private key permissions", "Paket Bağımlılığı", "-", "openssh-server kurulu", "SSH kurulu olmalı"],
    ["5.1.3-5.1.22", "SSH configuration rules", "Paket Bağımlılığı", "-", "openssh-server kurulu", "SSH kurulu olmalı"],

    # KERNEL MODÜL
    ["", "KERNEL MODÜL ZİNCİRİ", "", "", "", ""],
    ["3.2.1", "dccp disabled", "Bağımsız", "-", "-", "Diğerlerinden bağımsız"],
    ["3.2.2", "sctp disabled", "Bağımsız", "-", "-", "Diğerlerinden bağımsız"],
    ["3.2.3", "rds disabled", "Bağımsız", "-", "-", "Diğerlerinden bağımsız"],
    ["3.2.4", "tipc disabled", "Bağımsız", "-", "-", "Diğerlerinden bağımsız"],

    # SYSCTL
    ["", "SYSCTL PARAMETRELERİ", "", "", "", ""],
    ["1.5.1", "ASLR enabled", "Dosya Önceliği", "-", "60-kernel_sysctl.conf", "Aynı dosyada çakışma kontrolü"],
    ["3.3.1", "IP forwarding disabled", "Dosya Önceliği", "-", "60-netipv4_sysctl.conf", "Aynı dosyada çakışma kontrolü"],
    ["3.3.2", "Packet redirect disabled", "Dosya Önceliği", "-", "60-netipv4_sysctl.conf", "3.3.1 ile aynı dosya"],
    ["3.3.3-3.3.11", "Network sysctl params", "Dosya Önceliği", "-", "60-netipv*.conf", "Parametre çakışması kontrolü"],

    # SERVİS
    ["", "SERVİS BAĞIMLILIKLARI", "", "", "", ""],
    ["2.1.1", "autofs not in use", "Bağımsız", "-", "-", "Servis yoksa zaten uyumlu"],
    ["2.1.2-2.1.22", "Various services disabled", "Bağımsız", "-", "-", "Her biri bağımsız kontrol"],
]

for row in data1:
    ws1.append(row)

# Grup başlıkları için stil
group_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
for row_num in range(2, ws1.max_row + 1):
    if ws1.cell(row=row_num, column=1).value == "":
        for col in range(1, 7):
            ws1.cell(row=row_num, column=col).fill = group_fill
            ws1.cell(row=row_num, column=col).font = Font(bold=True)

apply_data_style(ws1, 2, ws1.max_row)
auto_column_width(ws1)

# ============================================
# SHEET 2: Bağımlılık Tipleri
# ============================================
ws2 = wb.create_sheet("Bagimlilik Tipleri")

headers2 = ["Tip", "Sembol", "Aciklama", "Ornek"]
ws2.append(headers2)
apply_header_style(ws2, 1, header_fill_green)

data2 = [
    ["Sıralı", "→", "A tamamlanmadan B uygulanamaz", "4.2.1 → 4.2.3"],
    ["Karşılıklı Dışlama", "⊕", "Sadece biri seçilebilir", "4.2.x ⊕ 4.3.x ⊕ 4.4.x"],
    ["Koşullu", "?", "Paket/servis varsa uygulanır", "1.7.2 ? (gdm3 kurulu)"],
    ["Ters Koşullu", "!", "Paket/servis yoksa uygulanır", "1.7.1 ! (gdm3 yok)"],
    ["Paket Bağımlılığı", "📦", "Paket kurulu olmalı", "5.1.x 📦 openssh-server"],
    ["Dosya Önceliği", "#", "Aynı dosyada sonraki geçerli", "sysctl.d/60-*.conf"],
    ["Bağımsız", "○", "Herhangi bir bağımlılık yok", "3.2.1, 3.2.2"],
    ["Ön Koşul", "⬆", "Tüm alt kuralların temeli", "5.3.1.1 (PAM version)"],
]

for row in data2:
    ws2.append(row)

apply_data_style(ws2, 2, ws2.max_row)
auto_column_width(ws2)

# ============================================
# SHEET 3: Kritik Uygulama Sırası
# ============================================
ws3 = wb.create_sheet("Kritik Uygulama Sirasi")

headers3 = ["Sira", "Grup", "Kurallar", "Neden?"]
ws3.append(headers3)
apply_header_style(ws3, 1, header_fill_orange)

data3 = [
    ["1", "Seçim", "4.1.1", "Firewall kararı önce verilmeli"],
    ["2", "Kurulum", "4.2.1 veya nftables paketi", "Seçime göre kurulum"],
    ["3", "Temizlik", "4.2.2, 4.3.2", "Çakışan paketler kaldırılmalı"],
    ["4", "Kurallar", "4.2.4-6 veya 4.3.7", "Bağlantı kuralları"],
    ["5", "Policy", "4.2.7 veya 4.3.8", "Default deny EN SON"],
    ["6", "Servis", "4.2.3 veya 4.3.9", "Kurallar hazır, servis başlat"],
]

for row in data3:
    ws3.append(row)

apply_data_style(ws3, 2, ws3.max_row)
auto_column_width(ws3)

# ============================================
# SHEET 4: Karşılıklı Dışlama Grupları
# ============================================
ws4 = wb.create_sheet("Karsilikli Dislama")

headers4 = ["Grup", "Secim", "Uygulanacak Kurallar", "Uygulanmayacak Kurallar"]
ws4.append(headers4)
apply_header_style(ws4, 1, header_fill_purple)

data4 = [
    ["Firewall", "UFW", "4.2.1 → 4.2.7", "4.3.x, 4.4.x"],
    ["Firewall", "nftables", "4.3.2 → 4.3.9", "4.2.x, 4.4.x"],
    ["Firewall", "iptables", "4.4.1.x", "4.2.x, 4.3.x"],
    ["GDM", "Server (GDM Kaldır)", "1.7.1", "1.7.2 - 1.7.10"],
    ["GDM", "Desktop (GDM Koru)", "1.7.2 - 1.7.10", "1.7.1"],
]

for row in data4:
    ws4.append(row)

apply_data_style(ws4, 2, ws4.max_row)
auto_column_width(ws4)

# ============================================
# SHEET 5: Firewall Detay
# ============================================
ws5 = wb.create_sheet("Firewall Bagimliliklari")

headers5 = ["Kural ID", "Kural Adi", "Onceki Kural", "Sonraki Kural", "Kritik Not"]
ws5.append(headers5)
apply_header_style(ws5, 1)

data5 = [
    ["4.1.1", "Single firewall choice", "-", "4.2.1 veya 4.3.2 veya 4.4.1", "KARAR NOKTASI"],
    ["", "--- UFW Zinciri ---", "", "", ""],
    ["4.2.1", "ufw installed", "4.1.1", "4.2.2, 4.2.3", "İlk adım"],
    ["4.2.2", "iptables-persistent removed", "4.2.1", "-", "Çakışma önleme"],
    ["4.2.3", "ufw service enabled", "4.2.1", "4.2.4, 4.2.5, 4.2.6", "Servis başlatma"],
    ["4.2.4", "loopback configured", "4.2.3", "4.2.7", "Temel kural"],
    ["4.2.5", "outbound configured", "4.2.3", "4.2.7", "Temel kural"],
    ["4.2.6", "firewall rules exist", "4.2.3", "4.2.7", "Port kuralları"],
    ["4.2.7", "default deny policy", "4.2.4, 4.2.5, 4.2.6", "-", "EN SON - yoksa kilitlenme!"],
    ["", "--- nftables Zinciri ---", "", "", ""],
    ["4.3.2", "ufw disabled", "4.1.1", "4.3.7", "UFW kapatılmalı"],
    ["4.3.7", "connections configured", "4.3.2", "4.3.8", "Bağlantı kuralları"],
    ["4.3.8", "default deny policy", "4.3.7", "4.3.9", "EN SON"],
    ["4.3.9", "nftables service enabled", "4.3.8", "-", "Servis başlatma"],
]

for row in data5:
    ws5.append(row)

apply_data_style(ws5, 2, ws5.max_row)
auto_column_width(ws5)

# ============================================
# SHEET 6: GDM Detay
# ============================================
ws6 = wb.create_sheet("GDM Bagimliliklari")

headers6 = ["Kural ID", "Kural Adi", "Kosul", "Profile", "GDM Gerekli mi?"]
ws6.append(headers6)
apply_header_style(ws6, 1, header_fill_green)

data6 = [
    ["1.7.1", "GDM is removed", "Server ortamı", "Level 2 - Server", "HAYIR (kaldırılıyor)"],
    ["1.7.2", "GDM login banner", "gdm3 kurulu", "Level 1 - Server/Workstation", "EVET"],
    ["1.7.3", "GDM disable-user-list", "gdm3 kurulu", "Level 1 - Server/Workstation", "EVET"],
    ["1.7.4", "GDM screen locks idle", "gdm3 + gnome kurulu", "Level 1 - Server/Workstation", "EVET + GNOME"],
    ["1.7.5", "GDM screen locks override", "1.7.4 uygulandı", "Level 1 - Server/Workstation", "EVET"],
    ["1.7.6", "GDM auto mount disabled", "gdm3 kurulu", "Level 1 - Server/Workstation", "EVET"],
    ["1.7.7", "GDM autorun-never", "gdm3 kurulu", "Level 1 - Server/Workstation", "EVET"],
    ["1.7.8", "GDM autorun override", "1.7.7 uygulandı", "Level 1 - Server/Workstation", "EVET"],
    ["1.7.9", "GDM disable-print-setup", "gdm3 kurulu", "Level 2 - Server/Workstation", "EVET"],
    ["1.7.10", "GDM XDMCP disabled", "gdm3 kurulu", "Level 1 - Server/Workstation", "EVET"],
]

for row in data6:
    ws6.append(row)

apply_data_style(ws6, 2, ws6.max_row)
auto_column_width(ws6)

# ============================================
# SHEET 7: PAM Detay
# ============================================
ws7 = wb.create_sheet("PAM Bagimliliklari")

headers7 = ["Kural ID", "Kural Adi", "Bagimlilik", "Konfigurasyon Dosyasi", "Not"]
ws7.append(headers7)
apply_header_style(ws7, 1, header_fill_orange)

data7 = [
    ["5.3.1.1", "PAM latest version", "-", "-", "TEMEL - PAM ≥1.5.3-5"],
    ["5.3.2.1", "pam_unix enabled", "5.3.1.1", "/etc/pam.d/common-*", "Temel auth modülü"],
    ["5.3.2.2", "pam_faillock enabled", "5.3.1.1", "/etc/pam.d/common-auth", "Hesap kilitleme"],
    ["5.3.2.3", "pam_pwquality enabled", "5.3.1.1", "/etc/pam.d/common-password", "Şifre kalitesi"],
    ["5.3.2.4", "pam_pwhistory enabled", "5.3.1.1", "/etc/pam.d/common-password", "Şifre geçmişi"],
    ["5.3.3.1.1", "faillock deny", "5.3.2.2", "/etc/security/faillock.conf", "deny = 5"],
    ["5.3.3.1.2", "faillock unlock_time", "5.3.2.2", "/etc/security/faillock.conf", "unlock_time = 900"],
    ["5.3.3.1.3", "faillock even_deny_root", "5.3.2.2", "/etc/security/faillock.conf", "root da kilitleniyor"],
    ["5.3.3.2.1", "pwquality minlen", "5.3.2.3", "/etc/security/pwquality.conf", "minlen = 14"],
    ["5.3.3.2.2", "pwquality minclass", "5.3.2.3", "/etc/security/pwquality.conf", "minclass = 4"],
    ["5.3.3.3.1", "pwhistory remember", "5.3.2.4", "/etc/security/pwhistory.conf", "remember = 24"],
    ["5.3.3.4.1", "pam_unix nullok", "5.3.2.1", "/etc/pam.d/common-password", "nullok olmamalı"],
]

for row in data7:
    ws7.append(row)

apply_data_style(ws7, 2, ws7.max_row)
auto_column_width(ws7)

# ============================================
# Dosyayı kaydet
# ============================================
output_path = r"c:\Users\cagri\OneDrive\Belgeler\GitHub\sh-bitirme-proje\docs\Kural_Bagimliliklari.xlsx"
wb.save(output_path)
print(f"Excel file created: {output_path}")
